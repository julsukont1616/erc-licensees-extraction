#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to split multi-fuel entries into separate rows with divided capacity.
Each row will have only ONE main fuel type (cleaned, without {}).
Technology details are moved to a separate column.
Capacity values (MW and kVA) are divided by the number of fuel types.
"""

import csv
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def parse_fuel_type(fuel_string):
    """
    Parse fuel type string and extract main fuel type and technology detail.

    Returns:
        tuple: (main_fuel_clean, tech_detail)
    """
    if not fuel_string or fuel_string.strip() == '':
        return None, None

    fuel_string = fuel_string.strip()

    # Extract technology detail in curly braces {}
    tech_detail = None
    curly_match = re.search(r'\{([^}]+)\}', fuel_string)
    if curly_match:
        tech_detail = curly_match.group(1)
        # Remove the curly braces part from main fuel string
        main_fuel = re.sub(r'\{[^}]+\}', '', fuel_string).strip()
    else:
        main_fuel = fuel_string

    return main_fuel, tech_detail

def split_fuels_smart(fuel_string):
    """
    Smart split that handles commas outside of curly braces.
    Returns list of individual fuel strings (with their tech details attached).
    """
    if not fuel_string or ',' not in fuel_string:
        return [fuel_string] if fuel_string else []

    parts = []
    depth = 0
    current_part = []

    for char in fuel_string:
        if char == '{':
            depth += 1
            current_part.append(char)
        elif char == '}':
            depth -= 1
            current_part.append(char)
        elif char == ',' and depth == 0:
            # This is a separator comma
            parts.append(''.join(current_part).strip())
            current_part = []
        else:
            current_part.append(char)

    # Don't forget the last part
    if current_part:
        parts.append(''.join(current_part).strip())

    return [p for p in parts if p]

def parse_number(value_str):
    """Parse a number from string, removing commas and handling empty values."""
    if not value_str or value_str.strip() == '':
        return 0.0
    # Remove commas and convert to float
    try:
        return float(str(value_str).replace(',', ''))
    except:
        return 0.0

def format_number(value, decimals=3):
    """Format a number with specified decimals and comma separators."""
    if value == 0:
        return ''
    return f"{value:,.{decimals}f}"

def main():
    input_csv = 'RadGridExport.csv'
    output_excel = 'RadGridExport_split_by_fuel.xlsx'

    print("Reading CSV file and analyzing structure...")

    # First pass: read the CSV to understand structure
    with open(input_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        original_header = next(reader)

    print(f"Original columns: {len(original_header)}")
    print("\nOriginal header structure:")
    for idx, col in enumerate(original_header):
        print(f"  Col {idx}: {col}")

    # Create new header with technology column inserted
    # Insert "เทคโนโลยี/รายละเอียดเชื้อเพลิงหลัก" after main fuel column (index 6)
    new_header = original_header[:7].copy()
    new_header.insert(7, "เทคโนโลยี/รายละเอียดเชื้อเพลิงหลัก (Technology Detail)")
    new_header.extend(original_header[7:])

    print(f"\nNew columns: {len(new_header)}")
    print("\nNew header structure:")
    for idx, col in enumerate(new_header):
        print(f"  Col {idx}: {col}")

    # Create Excel workbook
    print("\nCreating Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Electricity Licenses"

    # Write header to Excel with formatting
    for col_idx, header_val in enumerate(new_header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_val)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Column indices (0-based for array, 1-based for Excel)
    main_fuel_col_idx = 6      # Array index for main fuel in original CSV
    new_tech_col_idx = 7       # Array index for NEW tech column (after insertion)
    mw_col_idx = 8             # Original MW column in CSV (now will be at index 9 in new structure)
    kva_col_idx = 9            # Original kVA column in CSV (now will be at index 10 in new structure)

    # Read and process CSV
    with open(input_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)  # Skip header

        excel_row = 2
        total_csv_rows = 0
        total_split_rows = 0
        rows_with_splits = 0

        for csv_row_num, row in enumerate(reader, start=2):
            if len(row) <= main_fuel_col_idx:
                continue

            total_csv_rows += 1
            main_fuel_string = row[main_fuel_col_idx]

            # Split the main fuel into individual fuels with their tech details
            fuel_parts = split_fuels_smart(main_fuel_string)

            # Parse each fuel part to separate fuel from tech
            parsed_fuels = []
            for fuel_part in fuel_parts:
                fuel_clean, tech = parse_fuel_type(fuel_part)
                if fuel_clean:
                    parsed_fuels.append((fuel_clean, tech))

            # If no fuels parsed, use original (shouldn't happen)
            if not parsed_fuels:
                fuel_clean, tech = parse_fuel_type(main_fuel_string)
                parsed_fuels = [(fuel_clean or main_fuel_string, tech)]

            num_fuels = len(parsed_fuels)

            # Parse capacity values from original positions
            mw_value = parse_number(row[mw_col_idx]) if len(row) > mw_col_idx else 0.0
            kva_value = parse_number(row[kva_col_idx]) if len(row) > kva_col_idx else 0.0

            # Divide capacity by number of fuels
            divided_mw = mw_value / num_fuels if num_fuels > 0 and mw_value > 0 else mw_value
            divided_kva = kva_value / num_fuels if num_fuels > 0 and kva_value > 0 else kva_value

            if num_fuels > 1:
                rows_with_splits += 1
                total_split_rows += num_fuels

            # Create a row for each fuel type
            for fuel_idx, (fuel_clean, tech) in enumerate(parsed_fuels):
                # Build new row with inserted tech column
                new_row = []

                # Copy columns 0-6 (up to and including main fuel)
                for i in range(7):
                    if i == main_fuel_col_idx:
                        # Replace main fuel with CLEAN fuel name only
                        new_row.append(fuel_clean)
                    else:
                        new_row.append(row[i] if i < len(row) else '')

                # Insert technology detail in new column 7
                new_row.append(tech or '')

                # Copy remaining columns (supplementary fuel onwards)
                for i in range(7, len(row)):
                    if i == mw_col_idx:
                        # Replace MW with divided value
                        new_row.append(format_number(divided_mw, 3))
                    elif i == kva_col_idx:
                        # Replace kVA with divided value
                        new_row.append(format_number(divided_kva, 2))
                    else:
                        new_row.append(row[i])

                # Ensure row has enough columns
                while len(new_row) < len(new_header):
                    new_row.append('')

                # Write to Excel
                for col_idx, cell_value in enumerate(new_row, start=1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=cell_value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                    # Highlight split rows slightly
                    if num_fuels > 1 and fuel_idx > 0:
                        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

                excel_row += 1

            if csv_row_num % 100 == 0:
                print(f"  Processed {csv_row_num} CSV rows...")

    # Adjust column widths
    print("Adjusting column widths...")
    column_widths = {
        1: 40,   # Licensee name
        2: 50,   # Facility name
        3: 15,   # Province
        4: 20,   # Regional office
        5: 20,   # License number
        6: 20,   # Issue date
        7: 45,   # Main fuel (cleaned)
        8: 50,   # Technology detail (NEW)
        9: 35,   # Supplementary fuel
        10: 15,  # MW
        11: 15,  # kVA
        12: 20,  # COD date
    }

    for col_idx, width in column_widths.items():
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = width

    # Set default width for remaining columns
    for col_idx in range(len(column_widths) + 1, len(new_header) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 12

    # Freeze top row and first column
    ws.freeze_panes = 'B2'

    # Save Excel file
    print(f"\nSaving Excel file: {output_excel}")
    wb.save(output_excel)

    print("\n" + "="*100)
    print("SUMMARY")
    print("="*100)
    print(f"Original CSV rows:                    {total_csv_rows}")
    print(f"Excel output rows:                    {excel_row - 2}")
    print(f"Rows that had multiple fuels:         {rows_with_splits}")
    print(f"Total rows created from splits:       {total_split_rows}")
    print(f"Additional rows created:              {(excel_row - 2) - total_csv_rows}")
    print(f"\nNew column added:                     'เทคโนโลยี/รายละเอียดเชื้อเพลิงหลัก'")
    print(f"Position:                             After main fuel column (col 8)")
    print(f"\nOutput file:                          {output_excel}")
    print("="*100)

    # Run verification
    print("\n" + "="*100)
    print("VERIFICATION: Checking for {} in main fuel column")
    print("="*100)

    error_count = 0
    for row_idx in range(2, min(excel_row, ws.max_row + 1)):
        main_fuel_cell = ws.cell(row=row_idx, column=7).value  # Column 7 is main fuel
        if main_fuel_cell and '{' in str(main_fuel_cell):
            if error_count == 0:
                print(f"\nERROR: Found {{}} in main fuel column:")
            print(f"  Row {row_idx}: {main_fuel_cell}")
            error_count += 1
            if error_count >= 5:
                print(f"  ... and {error_count - 5} more")
                break

    if error_count == 0:
        print("\n✓ SUCCESS: No {} found in main fuel column! All entries are clean.")
    else:
        print(f"\n✗ FAILED: Found {error_count} rows with {{}} in main fuel column")

    print("="*100)

if __name__ == '__main__':
    main()
